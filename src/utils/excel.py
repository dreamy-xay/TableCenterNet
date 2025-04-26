#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Description:
Version:
Autor: dreamy-xay
Date: 2024-10-22 14:49:10
LastEditors: dreamy-xay
LastEditTime: 2024-10-22 14:49:13
"""
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment


def format_excel(
    xlsx_path,
    offset=(1, 1),
    row_width_extend=2,
    row_height=20,
    alignment=("center", "center"),
    side_config={"border_style": "thin", "color": "000000"},
):
    # Load Excel files using openpyxl
    wb = load_workbook(xlsx_path)

    # Make some customizations to each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]  # Get each worksheet object

        # Set a solid border (skip the leftmost column and the top side row)
        border = Border(left=Side(**side_config), right=Side(**side_config), top=Side(**side_config), bottom=Side(**side_config))

        # Add solid borders for all cells (but excluding the leftmost and uppermost columns)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = border

        # Offset down and to the right (insert empty rows and columns)
        for _ in range(offset[0]):  # Offset downward
            ws.insert_rows(1)
        for _ in range(offset[1]):  # Offset to the right
            ws.insert_cols(1)

        # Automatically adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get column letters
            for cell in col:
                if cell.value:
                    try:
                        max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
            if max_length != 0:
                adjusted_width = max_length + row_width_extend  # Add 2 for extra space
                ws.column_dimensions[column].width = adjusted_width

        # Automatically adjust the row height
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = row_height

        # Set all cell contents to be centered
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal=alignment[0], vertical=alignment[1])

    # Save the modified Excel file
    wb.save(xlsx_path)
